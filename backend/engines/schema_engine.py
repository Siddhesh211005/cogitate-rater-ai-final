import openpyxl
import os


# ── Main calculate function ───────────────────────────────────
# Called by both /api/schema/calculate and /api/raters/{slug}/calculate
def calculate(rater: dict, inputs: dict) -> dict:
    from engines.excel_engine import calculate as excel_calculate

    outputs, _meta = excel_calculate(rater, inputs)
    return outputs


# ── Core calculation logic ────────────────────────────────────
def calculate_from_file(filepath: str, config: dict, inputs: dict) -> dict:
    # Schema engine now reuses the hardened Excel calculation path so both
    # engines produce consistent values across test-calculate and live rating.
    from engines.excel_engine import calculate_from_file as excel_calculate_from_file
    outputs, _meta = excel_calculate_from_file(filepath, config, inputs)
    return outputs


# ── Evaluate using Python formulas library ────────────────────
def _evaluate_with_formulas(filepath: str, config: dict) -> dict:
    try:
        import formulas
        xl_model = formulas.ExcelModel().loads(filepath).finish()
        xl_model.calculate()

        outputs = {}
        sheet_name = config.get("sheet", "Rater")

        for field_def in config.get("outputs", []):
            field = field_def.get("field")
            cell  = field_def.get("cell")
            if not field or not cell:
                continue
            try:
                val = _get_formula_cell(xl_model, filepath, sheet_name, cell)

                if val is not None:
                    result = val.value
                    if hasattr(result, '__iter__') and not isinstance(result, str):
                        try:
                            result = next(iter(result.flat))
                        except Exception:
                            result = None
                    outputs[field] = _serialize(result)
                else:
                    outputs[field] = None
            except Exception:
                outputs[field] = None

        return outputs

    except ImportError:
        raise Exception("formulas library not installed")


def _evaluate_with_com(filepath: str, config: dict) -> dict:
    from engines.excel_engine import _calculate_with_com

    return _calculate_with_com(filepath, config)


def _get_formula_cell(xl_model, filepath: str, sheet_name: str, cell: str):
    for ref in _build_reference_candidates(filepath, sheet_name, cell):
        val = xl_model.cells.get(ref)
        if val is not None:
            return val

    # Fallback for workbooks where formulas normalizes refs differently
    target_cell = cell.upper()
    target_sheet = sheet_name.upper()
    loose_match = None

    for ref, val in xl_model.cells.items():
        normalized = _normalize_ref(ref)
        if "!" not in normalized:
            continue
        left, right = normalized.rsplit("!", 1)
        sheet_part = left.split("]")[-1]

        if right == target_cell and sheet_part == target_sheet:
            return val
        if right == target_cell and loose_match is None:
            loose_match = val

    return loose_match


def _build_reference_candidates(filepath: str, sheet_name: str, cell: str) -> list[str]:
    filename = os.path.basename(filepath)
    sheet_variants = [sheet_name, sheet_name.upper(), sheet_name.lower()]
    refs = []
    for sheet in sheet_variants:
        refs.extend([
            f"'{sheet}'!{cell.upper()}",
            f"{sheet}!{cell.upper()}",
            f"'[{filename}]{sheet}'!{cell.upper()}",
            f"[{filename}]{sheet}!{cell.upper()}",
        ])
    return list(dict.fromkeys(refs))


def _normalize_ref(ref: str) -> str:
    return str(ref).replace("'", "").strip().upper()


# ── Fallback: read calculated values with openpyxl ───────────
def _evaluate_with_openpyxl(filepath: str, config: dict) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheet_name = config.get("sheet")
    if not sheet_name or sheet_name not in wb.sheetnames:
        sheet_name = next(
            (s for s in wb.sheetnames if not s.startswith("_")),
            wb.sheetnames[0]
        )

    ws = wb[sheet_name]
    outputs = {}

    output_fields = config.get("outputs", [])
    for field_def in output_fields:
        field = field_def.get("field")
        cell  = field_def.get("cell")
        if not field or not cell:
            continue
        try:
            outputs[field] = _serialize(ws[cell].value)
        except Exception:
            outputs[field] = None

    wb.close()
    return outputs


# ── Helpers ───────────────────────────────────────────────────
def _cast_value(value, ftype: str):
    try:
        if ftype == "number":
            if isinstance(value, str):
                cleaned = value.strip().replace(",", "").replace("$", "")
                if cleaned.startswith("(") and cleaned.endswith(")"):
                    cleaned = f"-{cleaned[1:-1]}"
                return float(cleaned)
            return float(value)
        elif ftype == "dropdown":
            return str(value)
        else:
            return str(value)
    except Exception:
        return value


def _serialize(value):
    # formulas library can return an Empty token object
    # which stringifies as "empty"; normalize to None.
    if value is not None and str(value).strip().lower() in {"empty", "none", "null", ""}:
        return None
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 4)
    return str(value)


def _outputs_empty(outputs: dict | None) -> bool:
    if not outputs:
        return True
    return all(v is None for v in outputs.values())
