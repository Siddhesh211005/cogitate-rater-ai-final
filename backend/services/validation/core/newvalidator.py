from __future__ import annotations

from backend.services.validation.input.input_validator import validate_inputs
from backend.services.validation.input.cross_field_validator import validate_cross_fields
from backend.services.validation.rules.rule_engine import run_business_rules
from backend.services.validation.compliance.compliance_engine import check_compliance
from backend.services.validation.explainability.explainability_engine import generate_explanation
from backend.services.validation.profiling.data_profiling import profile_data
from backend.services.validation.profiling.schema_drift import detect_schema_drift
from backend.services.validation.comparison.rater_comparator import compare_raters
from backend.services.validation.scoring.severity_classifier import classify_issues
from backend.services.validation.scoring.scoring_engine import compute_risk_score

import io
import re
from collections import defaultdict, Counter
from typing import Dict, Any

import openpyxl


SUPPORTED_FUNCTIONS = {
    "SUM", "ROUND", "IF", "VLOOKUP", "INDEX", "MATCH", "MIN", "MAX", "AVERAGE"
}

FUNC_REGEX = re.compile(r"([A-Z]+)\(")


def _build_rule_schema(schema: Any) -> Dict[str, Dict[str, Any]] | None:

    if not isinstance(schema, dict):
        return None

    inputs = schema.get("inputs")
    if not isinstance(inputs, list):
        return None

    rules: Dict[str, Dict[str, Any]] = {}
    for field in inputs:
        if not isinstance(field, dict):
            continue
        name = field.get("field") or field.get("name") or field.get("key")
        if not name:
            continue
        rule: Dict[str, Any] = {}
        if "type" in field:
            rule["type"] = field.get("type")
        if "min" in field:
            rule["min"] = field.get("min")
        if "max" in field:
            rule["max"] = field.get("max")
        if "required" in field:
            rule["required"] = field.get("required")
        rules[name] = rule

    return rules or None


def _iter_schema_cells(schema: Any) -> list[str]:
    if not schema:
        return []

    if isinstance(schema, dict):
        inputs = schema.get("inputs")
        if isinstance(inputs, list):
            return [str(item.get("cell", "")).replace("$", "") for item in inputs if isinstance(item, dict)]
        return []

    if hasattr(schema, "input_fields"):
        return [str(field.cell_ref).replace("$", "") for field in schema.input_fields]

    return []


def _empty_validation_payload(reason: str | None = None) -> Dict[str, Any]:
    info = []
    if reason:
        info.append(reason)
    return {
        "errors": [],
        "warnings": [],
        "info": info,
    }


# ======================================================
# WORKBOOK SCAN
# ======================================================

def scan_workbook(workbook):

    worksheets = len(workbook.sheetnames)

    total_cells = 0
    formula_cells = 0
    formulas = []

    for sheet in workbook.sheetnames:

        ws = workbook[sheet]

        for row in ws.iter_rows():

            for cell in row:

                total_cells += 1

                if isinstance(cell.value, str) and cell.value.startswith("="):

                    formula_cells += 1
                    formulas.append(cell.value)

    return worksheets, total_cells, formula_cells, formulas


# ======================================================
# FORMULA ANALYSIS
# ======================================================

def analyze_formulas(formulas):

    parsed = 0
    unsupported = 0

    unsupported_functions = Counter()

    for f in formulas:

        funcs = FUNC_REGEX.findall(f.upper())

        bad = [fn for fn in funcs if fn not in SUPPORTED_FUNCTIONS]

        if bad:

            unsupported += 1

            for fn in bad:

                unsupported_functions[fn] += 1

        else:

            parsed += 1

    return parsed, unsupported, unsupported_functions


# ======================================================
# DEPENDENCY GRAPH
# ======================================================

def build_dependency_graph(workbook):

    graph = defaultdict(set)

    pattern = re.compile(r"([A-Za-z_]+!\$?[A-Z]+\$?[0-9]+|\$?[A-Z]+\$?[0-9]+)")

    for sheet in workbook.sheetnames:

        ws = workbook[sheet]

        for row in ws.iter_rows():

            for cell in row:

                if isinstance(cell.value, str) and cell.value.startswith("="):

                    formula = cell.value

                    matches = pattern.findall(formula)

                    node = f"{sheet}!{cell.coordinate}"

                    for m in matches:

                        graph[node].add(m)

    return graph


# ======================================================
# ENTERPRISE VALIDATOR
# ======================================================

def enterprise_validate(inputs, file_content, schema=None):

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=False)

        graph = build_dependency_graph(wb)

        worksheets, total_cells, formula_cells, formulas = scan_workbook(wb)

        parsed, unsupported, unsupported_functions = analyze_formulas(formulas)

        wb.close()

        input_validation = validate_inputs(inputs, schema)
        cross_validation = validate_cross_fields(inputs)
        rules = run_business_rules(inputs)
        compliance = check_compliance(inputs)
        explanation = generate_explanation(inputs, formulas)
        profiling = profile_data(inputs)
        schema_drift = detect_schema_drift(schema)
        comparison = compare_raters({}, [])
    except Exception as exc:
        payload = _empty_validation_payload("validation failed")
        payload["errors"].append(str(exc))
        return payload

    all_issues = []
    all_issues += input_validation.get("errors", [])
    all_issues += cross_validation.get("issues", [])
    all_issues += compliance.get("compliance_issues", [])

    severity = classify_issues(all_issues)

    risk = compute_risk_score({
        "errors": severity.get("critical", []),
        "warnings": severity.get("warning", [])
    })

    # "input_validation": input_validation,
    # "cross_field_validation": cross_validation,
    # "business_rules": rules,
    # "compliance": compliance,
    # "explainability": explanation,
    # "data_profiling": profiling,
    # "schema_drift": schema_drift,
    # "comparison": comparison,
    # "severity": severity,
    # "risk_score": risk,

    # ======================================================
    # PREMIUM DRIVER ANALYSIS
    # ======================================================

    inputs_affecting_premium = 0
    unused_inputs = 0

    if schema:
        for ref in _iter_schema_cells(schema):
            if not ref:
                continue

            used = False

            for f in formulas:

                if ref in f:

                    used = True
                    break

            if used:
                inputs_affecting_premium += 1
            else:
                unused_inputs += 1


    # ======================================================
    # QUALITY SCORE
    # ======================================================

    score = 100 - unsupported - (unused_inputs * 2)

    if score < 0:
        score = 0


    return {

        "errors": list(input_validation.get("errors", []) or []),
        "warnings": list(input_validation.get("warnings", []) or []),
        "info": [],

        "dependency_graph_nodes": len(graph),

        "premium_nodes": list(graph.keys())[:16],

        "lookup_tables": [],

        "formula_cycles": [],

        "workbook_analysis": {
            "worksheets": worksheets,
            "total_cells_scanned": total_cells,
            "formula_cells": formula_cells
        },

        "formula_extraction": {
            "detected": len(formulas),
            "parsed": parsed,
            "unsupported": unsupported
        },

        "unsupported_functions": dict(unsupported_functions),

        "premium_driver_analysis": {
            "inputs_affecting_premium": inputs_affecting_premium,
            "unused_inputs": unused_inputs
        },

        "dynamic_input_analysis": {
            "status": "NOT_DYNAMIC",
            "message": "Premium is calculated using default Excel inputs only."
        },

        "rater_quality_score": {
            "overall_score": score
        },

        "input_validation": input_validation,
        "cross_field_validation": cross_validation,
        "business_rules": rules,
        "compliance": compliance,
        "explainability": explanation,
        "data_profiling": profiling,
        "schema_drift": schema_drift,
        "comparison": comparison,
    }
