import re
import logging
from pathlib import Path
from typing import Any
from collections import Counter

import openpyxl

logger = logging.getLogger(__name__)

SCHEMA_SHEET_NAME = "_Schema"
CLASS_BUCKET_PATTERN = re.compile(r"^class\s+(?:[ivxlcdm]+|\d+)$", re.IGNORECASE)
BOUNDARY_SECTION_MARKERS = {"boundary", "parameter", "minimum", "maximum", "breach"}
# Pattern to detect data-code labels like '78PT10USD', '1MNS', etc.
_DATA_CODE_PATTERN = re.compile(r"^\d+[A-Za-z]")


def check_schema_sheet(filepath: str | Path) -> bool:
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        found = SCHEMA_SHEET_NAME in wb.sheetnames
        wb.close()
        return found
    except Exception:
        return False


def parse_schema(filepath: str | Path, *, require_schema_sheet: bool = False) -> dict[str, Any]:
    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=False)
    try:
        has_schema_sheet = SCHEMA_SHEET_NAME in wb.sheetnames
        if has_schema_sheet:
            config = _parse_from_schema_sheet(wb)
        else:
            if require_schema_sheet:
                raise ValueError(
                    "No '_Schema' sheet found in this workbook. The Excel engine requires a _Schema sheet."
                )
            config = _infer_from_workbook(wb)

        if not config.get("sheet"):
            config["sheet"] = _resolve_rater_sheet(wb)

        config = normalize_config(config)
        config["has_schema_sheet"] = has_schema_sheet
        config["sheets"] = wb.sheetnames
        return config
    finally:
        wb.close()


def _parse_options(options_raw: Any) -> list[Any]:
    if options_raw is None or str(options_raw).strip() == "":
        return []

    parsed = []
    for part in str(options_raw).split(";"):
        text = part.strip()
        if not text:
            continue
        try:
            parsed.append(int(text))
            continue
        except ValueError:
            pass
        try:
            parsed.append(float(text))
            continue
        except ValueError:
            pass
        parsed.append(text)
    return parsed


def _parse_default(default_raw: Any, field_type: str) -> Any:
    if default_raw is None or str(default_raw).strip() == "":
        return None
    if field_type != "number":
        return default_raw

    text = str(default_raw).strip()
    try:
        return float(text) if "." in text else int(text)
    except ValueError:
        return default_raw


def _parse_from_schema_sheet(wb) -> dict[str, Any]:
    ws = wb[SCHEMA_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("_Schema sheet is empty.")

    headers = [str(header).strip().lower() if header else "" for header in rows[0]]
    inputs: list[dict[str, Any]] = []
    outputs: list[dict[str, Any]] = []
    seen_fields: set[str] = set()
    schema_sheet = None

    for row in rows[1:]:
        if not any(row):
            continue

        entry = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            entry[header] = row[index] if index < len(row) else None

        field_raw = entry.get("field")
        cell = entry.get("cell")
        if not cell:
            continue

        cell = str(cell).strip()
        row_sheet = entry.get("sheet")
        if row_sheet and not schema_sheet:
            schema_sheet = str(row_sheet).strip()

        field_seed = field_raw or entry.get("label") or cell
        field = _unique_field_name(_sanitize_field_name(str(field_seed)), seen_fields)
        field_type = str(entry.get("type") or "text").strip()
        direction = str(entry.get("direction") or "input").strip().lower()
        label = str(entry.get("label") or field_raw or field).strip()
        group = str(entry.get("group") or "General").strip()
        options = _parse_options(entry.get("options"))
        default = _parse_default(entry.get("default"), field_type)

        item: dict[str, Any] = {
            "field": field,
            "cell": cell,
            "type": field_type,
            "label": label,
            "group": group,
        }

        if direction == "output":
            item["primary"] = _to_bool(entry.get("primary", False))
            if field == "premium":
                item["primary"] = True
            outputs.append(item)
        else:
            if options:
                item["options"] = options
                item["type"] = "dropdown"
            if default is not None:
                item["default"] = default
            inputs.append(item)

    if not inputs and not outputs:
        raise ValueError("_Schema sheet has no usable field definitions.")

    if outputs and not any(output.get("primary") for output in outputs):
        outputs[0]["primary"] = True

    config = {
        "sheet": schema_sheet or _resolve_rater_sheet(wb),
        "inputs": inputs,
        "outputs": outputs,
    }
    _inject_schedule_mode(config)
    return config


# ── Infer schema from workbook structure (Schema engine path) ──
def _infer_from_workbook(wb) -> dict:
    # Skip hidden/utility sheets and infer from the "best" candidate sheet.
    skip = {"_schema", "application", "pick list", "concatanate", "code"}
    candidate_sheets = [
        name for name in wb.sheetnames
        if name.lower() not in skip and not name.startswith("_")
    ]

    if not candidate_sheets:
        return {"inputs": [], "outputs": [], "sheet": None}

    best_sheet = None
    best_inputs = []
    best_outputs = []
    best_score = float("-inf")

    for sheet_name in candidate_sheets:
        ws = wb[sheet_name]
        inputs, outputs = _infer_fields_from_sheet(ws)
        score = _score_inferred_sheet(sheet_name, ws, inputs, outputs)
        if score > best_score:
            best_sheet = sheet_name
            best_inputs = inputs
            best_outputs = outputs
            best_score = score

    if not best_sheet:
        return {"inputs": [], "outputs": [], "sheet": None}

    return {
        "inputs": best_inputs,
        "outputs": best_outputs,
        "sheet": best_sheet,
        "inferred": True,
    }


def _score_inferred_sheet(sheet_name: str, ws, inputs: list[dict], outputs: list[dict]) -> float:
    # Avoid selecting large lookup/rate tables over the actual user-facing
    # input sheet by combining quality signals rather than raw field count.
    inferred_count = len(inputs) + len(outputs)
    score = float(min(inferred_count, 120))

    normalized_name = _normalize_heading(sheet_name)
    if "input" in normalized_name and "output" in normalized_name:
        score += 300
    elif "input" in normalized_name:
        score += 140
    elif "output" in normalized_name:
        score += 90

    heading_hits = _count_heading_hits(ws)
    score += min(heading_hits, 6) * 25

    if not outputs:
        score -= 25

    labels = [str(item.get("label", "")).strip().lower() for item in inputs if str(item.get("label", "")).strip()]
    if labels:
        counts = Counter(labels)
        duplicate_labels = sum(count - 1 for count in counts.values() if count > 1)
        duplicate_ratio = duplicate_labels / len(labels)
        score -= duplicate_ratio * 180

    if inferred_count > 250:
        score -= 120
    if inferred_count > 800:
        score -= 160

    return score


def _count_heading_hits(ws, *, max_rows: int = 80, max_cols: int = 20) -> int:
    heading_hits = 0
    row_cap = min(max(ws.max_row or 1, 1), max_rows)
    col_cap = min(max(ws.max_column or 1, 1), max_cols)

    for row_num in range(1, row_cap + 1):
        for col_num in range(1, col_cap + 1):
            value = ws.cell(row=row_num, column=col_num).value
            if isinstance(value, str) and _is_heading_text(value):
                heading_hits += 1

    return heading_hits


def _infer_fields_from_sheet(ws) -> tuple[list[dict], list[dict]]:
    # Bound scanning to the likely active area; avoids noisy trailing columns/rows.
    max_col = min(max(ws.max_column or 1, 1), 40)
    max_row = min(max(ws.max_row or 1, 1), 500)

    inputs: list[dict] = []
    outputs: list[dict] = []
    seen_fields: set[str] = set()
    seen_value_cells: set[str] = set()
    seen_label_cells: set[str] = set()          # ← NEW: prevent same label cell being used twice
    seen_labels: set[str] = set()               # ← NEW: deduplicate by normalised label text

    # Pre-scan: identify column regions that belong to the boundary conditions
    # table so we can skip them entirely.
    boundary_columns = _detect_boundary_columns(ws, max_row, max_col)

    for row_num in range(1, max_row + 1):
        non_empty_cells = []
        for col_num in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if _has_value(cell.value):
                non_empty_cells.append(cell)

        if len(non_empty_cells) < 2:
            continue

        for label_cell in non_empty_cells:
            # Skip if this cell was already used as a label
            if label_cell.coordinate in seen_label_cells:
                continue

            # Skip cells already consumed as value cells (e.g. 'HKD' at D10)
            if label_cell.coordinate in seen_value_cells:
                continue

            # Skip cells in boundary-condition column regions
            if label_cell.column in boundary_columns:
                continue

            label = label_cell.value
            if not _is_label_candidate(label):
                continue

            label_str = str(label).strip()
            if _is_heading_text(label_str):
                continue

            # Skip data-code-like labels (e.g. '78PT10USD', '1MNS')
            if _is_data_code(label_str):
                continue

            # Deduplicate by normalised label text
            label_key = label_str.strip().lower()
            if label_key in seen_labels:
                continue

            value_cell = _find_value_cell_to_right(
                ws=ws,
                row_num=row_num,
                start_col=label_cell.column,
                max_col=max_col,
                seen_value_cells=seen_value_cells,
            )
            if value_cell is None:
                continue

            # Skip if the value cell is in a boundary column region
            if value_cell.column in boundary_columns:
                continue

            section_header = _find_section_header(ws, row_num, label_cell.column, max_col)
            if _is_excluded_section(section_header):
                continue

            # Also check if this row has boundary markers in neighbouring cells
            if _row_in_boundary_region(ws, row_num, label_cell.column, max_col):
                continue

            raw_value = value_cell.value
            field_name = _unique_field_name(_sanitize_field_name(label_str), seen_fields)
            is_output = _is_output_section(section_header) or _looks_like_formula(raw_value)
            is_number = _is_number_like(raw_value) or _looks_like_formula(raw_value)

            if is_output:
                outputs.append({
                    "field": field_name,
                    "cell": value_cell.coordinate,
                    "type": "number" if is_number else "text",
                    "label": label_str,
                    "group": "Results",
                    "default": None,
                    "primary": len(outputs) == 0,
                })
            else:
                inputs.append({
                    "field": field_name,
                    "cell": value_cell.coordinate,
                    "type": "number" if is_number else "text",
                    "label": label_str,
                    "group": "Rating Inputs",
                    "default": raw_value,
                    "options": [],
                })

            seen_value_cells.add(value_cell.coordinate)
            seen_label_cells.add(label_cell.coordinate)
            seen_labels.add(label_key)

    return inputs, outputs


def _find_value_cell_to_right(ws, row_num: int, start_col: int, max_col: int, seen_value_cells: set[str]):
    # Support organized layouts where labels and values are separated by one or more columns
    # (e.g. B label -> D value, F label -> H value).
    for col_num in range(start_col + 1, min(start_col + 4, max_col) + 1):
        candidate = ws.cell(row=row_num, column=col_num)
        if candidate.coordinate in seen_value_cells:
            continue
        if not _has_value(candidate.value):
            continue
        if _is_heading_text(str(candidate.value).strip()):
            continue
        return candidate
    return None


def _find_section_header(ws, row_num: int, col_num: int, max_col: int) -> str:
    best_text = ""
    best_priority = -1
    best_distance = 999

    for r in range(row_num - 1, max(0, row_num - 7), -1):
        distance = row_num - r
        for c in range(max(1, col_num - 2), min(max_col, col_num + 2) + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            text = v.strip()
            if _is_heading_text(text):
                priority = _heading_priority(text)
                if priority > best_priority or (priority == best_priority and distance < best_distance):
                    best_text = text
                    best_priority = priority
                    best_distance = distance

    return best_text


def _has_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def _is_label_candidate(value) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    if len(text) < 2:
        return False
    if text.startswith("="):
        return False
    if text.replace(".", "").replace("-", "").isdigit():
        return False
    return True


def _is_heading_text(value: str) -> bool:
    text = _normalize_heading(value)
    if not text:
        return False
    if text.startswith("please fill"):
        return True
    return text in {
        "output",
        "outputs",
        "result",
        "results",
        "input",
        "inputs",
        "boundary conditions",
        "parameter",
        "minimum",
        "maximum",
    }


def _heading_priority(value: str) -> int:
    text = _normalize_heading(value)
    if "boundary" in text:
        return 40
    if "output" in text or "result" in text:
        return 30
    if "input" in text or text.startswith("please fill"):
        return 20
    if text in BOUNDARY_SECTION_MARKERS:
        return 10
    return 0


def _normalize_heading(value: str) -> str:
    return re.sub(r"[^a-z0-9 ]+", "", value.strip().lower()).strip()


def _is_output_section(header: str) -> bool:
    h = (header or "").strip().lower()
    return "output" in h or "result" in h


def _is_excluded_section(header: str) -> bool:
    h = _normalize_heading(header or "")
    return any(marker in h for marker in BOUNDARY_SECTION_MARKERS)


def _detect_boundary_columns(ws, max_row: int, max_col: int) -> set[int]:
    """Pre-scan the first ~15 rows to find columns that belong to a boundary
    conditions table (headers like 'Parameter', 'Minimum', 'Maximum', 'Breach?').
    Returns a set of column numbers to exclude."""
    boundary_cols: set[int] = set()
    scan_rows = min(max_row, 15)
    for r in range(1, scan_rows + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and _normalize_heading(v) in BOUNDARY_SECTION_MARKERS:
                boundary_cols.add(c)
    return boundary_cols


def _row_in_boundary_region(ws, row_num: int, label_col: int, max_col: int) -> bool:
    """Check if any cell within ±2 columns of the label contains a boundary
    keyword, indicating the label is part of a boundary conditions table.
    Uses a narrow window to avoid blocking legitimate labels in distant
    column groups (e.g. output labels at col F when boundary is at col J)."""
    for c in range(max(1, label_col - 2), min(max_col, label_col + 2) + 1):
        v = ws.cell(row=row_num, column=c).value
        if isinstance(v, str) and _normalize_heading(v) in BOUNDARY_SECTION_MARKERS:
            return True
    return False


def _is_data_code(label: str) -> bool:
    """Detect concatenated data codes like '78PT10USD', '1MNS', '5FS'.
    These are reference codes from lookup tables, not user-facing field labels."""
    text = label.strip()
    if not text:
        return False
    # Starts with digit(s) followed by letters with no spaces → data code
    if _DATA_CODE_PATTERN.match(text) and " " not in text:
        return True
    return False


def _is_number_like(value) -> bool:
    if isinstance(value, (int, float)):
        return True
    if not isinstance(value, str):
        return False
    raw = value.strip().replace(",", "").replace("$", "")
    if not raw:
        return False
    if raw.startswith("(") and raw.endswith(")"):
        raw = "-" + raw[1:-1]
    if raw.endswith("%"):
        raw = raw[:-1]
    try:
        float(raw)
        return True
    except Exception:
        return False


def _resolve_rater_sheet(wb) -> str | None:
    for name in wb.sheetnames:
        if not name.startswith("_") and name.lower() != "application":
            return name
    return wb.sheetnames[0] if wb.sheetnames else None


def _sanitize_field_name(raw: str) -> str:
    base = (
        raw.lower()
        .replace(" ", "_")
        .replace("/", "_")
        .replace("-", "_")
        .replace("(", "")
        .replace(")", "")
        .replace(".", "")
        .replace(",", "")
        .replace("'", "")
        .replace("%", "pct")
        [:40]
        .strip("_")
    )
    return base or "field"


def _unique_field_name(base: str, seen: set[str]) -> str:
    candidate = base
    i = 2
    while candidate in seen:
        candidate = f"{base}_{i}"
        i += 1
    seen.add(candidate)
    return candidate


def _to_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def _looks_like_formula(value) -> bool:
    return isinstance(value, str) and value.strip().startswith("=")


def normalize_config(config: dict) -> dict:
    if not isinstance(config, dict):
        return config

    normalized = dict(config)
    inputs = normalized.get("inputs")
    if isinstance(inputs, list):
        normalized["inputs"] = _normalize_class_inputs(inputs)
    return normalized


def _normalize_class_inputs(inputs: list[dict]) -> list[dict]:
    normalized_inputs: list[dict] = []
    class_options: list[str] = []
    class_field_index: int | None = None

    for field in inputs:
        field_copy = dict(field)
        label = str(field_copy.get("label", "")).strip()
        label_lower = label.lower()

        if label_lower == "class":
            if class_field_index is None:
                class_field_index = len(normalized_inputs)
                normalized_inputs.append(field_copy)
            continue

        if CLASS_BUCKET_PATTERN.match(label_lower):
            if label and label not in class_options:
                class_options.append(label)
            continue

        normalized_inputs.append(field_copy)

    if class_field_index is not None and class_options:
        class_field = normalized_inputs[class_field_index]

        existing_options = class_field.get("options") or []
        merged_options = list(existing_options)
        for option in class_options:
            if option not in merged_options:
                merged_options.append(option)

        class_field["type"] = "dropdown"
        class_field["options"] = merged_options

        default_value = class_field.get("default")
        default_str = str(default_value).strip() if default_value is not None else ""
        if (
            not default_str
            or _looks_like_formula(default_value)
            or default_str not in merged_options
        ):
            class_field["default"] = merged_options[0]

    return normalized_inputs
# ── Auto-generate _Schema from workbook (no-schema fallback) ──
# Used when admin picks "Auto-generate Schema" option
# (Section 4 of context file — no-schema fallback UI)
def auto_generate_schema(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        rater_sheet = next(
            (name for name in wb.sheetnames if not name.startswith("_") and name.lower() != "application"),
            None,
        )
        if not rater_sheet:
            return {"inputs": [], "outputs": [], "auto_generated": True}

        ws = wb[rater_sheet]
        inputs = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                col = openpyxl.utils.get_column_letter(cell.column)
                if col != "B":
                    continue
                ref = f"{col}{cell.row}"
                value = cell.value
                inputs.append(
                    {
                        "field": f"field_{ref.lower()}",
                        "cell": ref,
                        "type": "number" if isinstance(value, (int, float)) else "text",
                        "label": f"Field {ref}",
                        "group": "Auto-Generated",
                        "default": value,
                    }
                )

        return {
            "inputs": inputs,
            "outputs": [],
            "sheet": rater_sheet,
            "auto_generated": True,
            "has_schema_sheet": False,
            "note": "Auto-generated schema. Review and edit field labels and outputs before saving.",
        }
    finally:
        wb.close()


def _inject_schedule_mode(config: dict[str, Any]) -> None:
    inputs = config.get("inputs") or []
    if not inputs:
        return

    row_map: dict[int, dict[str, dict[str, Any]]] = {}
    for item in inputs:
        cell = str(item.get("cell") or "").strip().upper()
        match = re.match(r"^([A-Z]+)(\d+)$", cell)
        if not match:
            continue
        col, row_text = match.groups()
        row_map.setdefault(int(row_text), {})[col] = item

    schedule_rows = sorted(row for row, cols in row_map.items() if "D" in cols)
    if not schedule_rows:
        return

    blocks = []
    start = schedule_rows[0]
    previous = schedule_rows[0]
    for row in schedule_rows[1:]:
        if row == previous + 1:
            previous = row
            continue
        blocks.append((start, previous))
        start = row
        previous = row
    blocks.append((start, previous))

    schedules = []
    for index, (row_start, row_end) in enumerate(blocks, 1):
        if row_end - row_start + 1 < 3:
            continue

        d_meta = row_map[row_start].get("D", {})
        e_meta = {}
        for row in range(row_start, row_end + 1):
            if "E" in row_map.get(row, {}):
                e_meta = row_map[row].get("E", {})
                break

        group_name = str(d_meta.get("group") or "Coverage")
        key_base = re.sub(r"[^a-z0-9]+", "_", group_name.lower()).strip("_")
        key = f"{key_base}_{row_start}_{row_end}" if key_base else f"schedule_{index}"

        schedules.append(
            {
                "key": key,
                "title": f"{group_name} ({row_start}-{row_end})",
                "rowStart": row_start,
                "rowEnd": row_end,
                "allowBlankRows": True,
                "minActiveRows": 1,
                "columns": [
                    {
                        "field": "location",
                        "column": "D",
                        "type": d_meta.get("type", "text"),
                        "label": str(d_meta.get("label") or "Location #"),
                    },
                    {
                        "field": "expiring_risk_limit",
                        "column": "E",
                        "type": e_meta.get("type", "number"),
                        "label": str(e_meta.get("label") or "Risk Limit"),
                    },
                ],
            }
        )

    if not schedules:
        return

    config["mode"] = "schedule"
    config["writeRules"] = {
        "clearUnusedRows": True,
        "emptyCellWrite": "blank",
        "rowActivePolicy": "any-non-empty-cell",
    }
    config["schedules"] = schedules
